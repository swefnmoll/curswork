import sqlite3 as sq
import openpyxl as xl

# Подключение к базе данных
connection = sq.connect('intonation.db')
cursor = connection.cursor()

def read_dictors():
    # Подключение к xlsx-файлу
    wb = xl.load_workbook('dictors.xlsx')
    sh = wb.active
    MAX_COLUMN = sh.max_column + 1
    MAX_ROW = sh.max_row + 1

    # Считывание данных из строки
    def read_row(row):
        dictor_data = []
        for column in range(1, MAX_COLUMN):
            dictor_data.append(sh.cell(row=row, column=column).value)
        return dictor_data

    # Запись данных
    def write_row_data(dictor_data):
        cursor.execute(f'''
        INSERT INTO dictors (name, gender, DOB, lang, settlement, education)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', tuple(dictor_data))
        connection.commit()

    # Считывание всех строк
    def read_full_xl():
        for row in range(1, MAX_ROW):
            write_row_data(read_row(row))

    read_full_xl()